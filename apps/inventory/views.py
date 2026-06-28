from rest_framework import viewsets, permissions
from .models import Branch, Category, Product, Stock, Purchase, Supplier, StockTransfer, PurchaseOrder, PurchaseOrderItem, Truck, TruckAllocation, StockAdjustment, GoodsReceivedNote, GRNItem, Driver, TruckMaintenance, DriverIssue
from .serializers import (
    BranchSerializer, CategorySerializer, ProductSerializer, 
    StockSerializer, PurchaseSerializer, SupplierSerializer, StockTransferSerializer,
    PurchaseOrderSerializer, PurchaseOrderItemSerializer, TruckSerializer, TruckAllocationSerializer, StockAdjustmentSerializer,
    GoodsReceivedNoteSerializer, GRNItemSerializer, DriverSerializer, TruckMaintenanceSerializer
)
import io
import csv
import openpyxl
from datetime import date
from decimal import Decimal
from django.db import transaction
from rest_framework import status
from rest_framework.response import Response
from django.db.models import Q, Sum
from rest_framework.decorators import action
from django.http import HttpResponse
from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin

from apps.users.permissions import IsStoreManager, IsStoreKeeper, IsStockController, IsAfisaUgavi

class BranchViewSet(viewsets.ModelViewSet):
    queryset = Branch.objects.all()
    serializer_class = BranchSerializer
    permission_classes = [permissions.DjangoModelPermissions]

class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [permissions.DjangoModelPermissions]

class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    permission_classes = [permissions.DjangoModelPermissions]
    filterset_fields = ['product_type', 'category']

    @action(detail=False, methods=['POST'], url_path='import')
    def import_products(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

        file_name = file.name
        headers = []
        rows = []
        imported_count = 0
        errors = []

        try:
            print(f"Starting import for file: {file_name}")
            if file_name.endswith('.csv'):
                decoded_file = file.read().decode('utf-8')
                io_string = io.StringIO(decoded_file)
                reader = csv.DictReader(io_string)
                rows = list(reader)
                headers = reader.fieldnames or []
            elif file_name.endswith(('.xlsx', '.xls')):
                wb = openpyxl.load_workbook(file, data_only=True)
                sheet = wb.active
                # Get headers and strip them
                headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
                for row_idx, row_data in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    row_dict = dict(zip(headers, row_data))
                    if any(row_dict.values()):  # Skip empty rows
                        rows.append(row_dict)
            else:
                return Response({"error": "Unsupported file format. Please upload CSV or Excel."}, status=status.HTTP_400_BAD_REQUEST)
            print(f"Read {len(rows)} rows from file. Headers: {headers}")
        except Exception as e:
            print(f"Error reading file: {str(e)}")
            return Response({"error": f"Failed to read file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            for i, row in enumerate(rows):
                try:
                    # Case-insensitive row lookup helper
                    def get_case_insensitive(d, key, default=None):
                        low_key = key.lower()
                        for k, v in d.items():
                            if k and k.lower() == low_key:
                                return v
                        return default

                    category_name = get_case_insensitive(row, 'Category')
                    if not category_name:
                        errors.append(f"Row {i+2}: Missing Category")
                        continue
                        
                    category, _ = Category.objects.get_or_create(name=str(category_name).strip())

                    sku = get_case_insensitive(row, 'SKU')
                    sku = str(sku).strip() if sku else None

                    product_name = get_case_insensitive(row, 'Name')
                    if not product_name:
                        errors.append(f"Row {i+2}: Missing Product Name")
                        continue

                    def get_val(key, default=''):
                        val = get_case_insensitive(row, key)
                        return val if val is not None else default

                    # If SKU exists, try to find by SKU. If not, try to find by Name + Category to prevent duplicates
                    product = None
                    created = False
                    
                    if sku:
                        product, created = Product.objects.get_or_create(
                            sku=sku,
                            defaults={
                                'name': str(product_name).strip(),
                                'category': category,
                                'product_type': str(get_val('Type', 'product')).lower().strip(),
                                'cost': float(str(get_val('Cost', 0) or 0).replace(',', '')),
                                'price': float(str(get_val('Price', 0) or 0).replace(',', '')),
                                'weight': float(str(get_val('Weight (kg)', 0) or 0).replace(',', '')),
                                'description': str(get_val('Description', '')).strip(),
                            }
                        )
                    else:
                        # No SKU provided, check by Name and Category
                        product, created = Product.objects.get_or_create(
                            name=str(product_name).strip(),
                            category=category,
                            defaults={
                                'product_type': str(get_val('Type', 'product')).lower().strip(),
                                'cost': float(str(get_val('Cost', 0) or 0).replace(',', '')),
                                'price': float(str(get_val('Price', 0) or 0).replace(',', '')),
                                'weight': float(str(get_val('Weight (kg)', 0) or 0).replace(',', '')),
                                'description': str(get_val('Description', '')).strip(),
                            }
                        )

                    # Handle Opening Stock (Additive)
                    try:
                        raw_opening = get_val('Opening Stock', 0)
                        raw_low = get_val('Low Stock Alert', 10)
                        opening_stock = int(float(str(raw_opening or 0).replace(',', '')))
                        low_stock = int(float(str(raw_low or 10).replace(',', '')))
                    except (ValueError, TypeError):
                        opening_stock = 0
                        low_stock = 10
                    
                    # Ensure Stock record exists in the Main Branch for 'product' types
                    product_type = str(get_val('Type', 'product')).lower().strip()
                    if product_type == 'product':
                        branch, _ = Branch.objects.get_or_create(name="Main Branch")
                        stock_obj, s_created = Stock.objects.get_or_create(
                            product=product,
                            branch=branch,
                            defaults={
                                'quantity': opening_stock,
                                'low_stock_threshold': low_stock
                            }
                        )
                        if not s_created:
                            # If stock record already existed, ADD to it
                            stock_obj.quantity += opening_stock
                            stock_obj.save()

                    imported_count += 1
                except Exception as e:
                    errors.append(f"Row {i+2}: {str(e)}")

        return Response({
            "message": f"Processed {len(rows)} rows, successfully imported/found {imported_count} products.",
            "errors": errors,
            "detected_headers": headers
        }, status=status.HTTP_201_CREATED if not errors else status.HTTP_207_MULTI_STATUS)

class StockViewSet(viewsets.ModelViewSet):
    queryset = Stock.objects.all()
    serializer_class = StockSerializer
    permission_classes = [permissions.DjangoModelPermissions]
    filterset_fields = ['branch', 'product']

class SupplierViewSet(viewsets.ModelViewSet):
    queryset = Supplier.objects.all()
    serializer_class = SupplierSerializer
    permission_classes = [permissions.DjangoModelPermissions]

class PurchaseViewSet(viewsets.ModelViewSet):
    queryset = Purchase.objects.all()
    serializer_class = PurchaseSerializer
    permission_classes = [permissions.DjangoModelPermissions]

    def create(self, request, *args, **kwargs):
        payload = request.data.copy()
        serializer = self.get_serializer(data=payload)
        serializer.is_valid(raise_exception=True)
        
        with transaction.atomic():
            # total_cost is computed (read-only in the serializer) so clients
            # only send quantity + unit_cost.
            total_cost = serializer.validated_data['quantity'] * serializer.validated_data['unit_cost']
            purchase = serializer.save(total_cost=total_cost)
            # Increase Stock
            stock, _ = Stock.objects.get_or_create(
                product=purchase.product,
                branch=purchase.branch,
                defaults={'quantity': 0}
            )
            stock.quantity += purchase.quantity
            stock.save()

        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

class StockTransferViewSet(viewsets.ModelViewSet):
    queryset = StockTransfer.objects.all()
    serializer_class = StockTransferSerializer
    permission_classes = [permissions.DjangoModelPermissions]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        transfer = serializer.save()

        # Update Stocks
        with transaction.atomic():
            # Decrease Source
            source_stock, _ = Stock.objects.get_or_create(product=transfer.product, branch=transfer.from_branch)
            source_stock.quantity -= transfer.quantity
            source_stock.save()

            # Increase Dest
            dest_stock, _ = Stock.objects.get_or_create(product=transfer.product, branch=transfer.to_branch)
            dest_stock.quantity += transfer.quantity
            dest_stock.save()

        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

class PurchaseOrderViewSet(viewsets.ModelViewSet):
    queryset = PurchaseOrder.objects.all()
    serializer_class = PurchaseOrderSerializer
    permission_classes = [permissions.IsAuthenticated, IsAfisaUgavi]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def destroy(self, request, *args, **kwargs):
        po = self.get_object()
        if po.status == 'received':
            return Response(
                {'detail': 'Cannot delete an order whose goods were already received (stock has been updated).'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['POST'])
    def receive(self, request, pk=None):
        """Confirm goods have arrived: increase stock for every line item and
        mark the order Received. Idempotent — stock is only added once."""
        po = self.get_object()
        if po.status == 'received':
            return Response({'detail': 'This order is already marked as received.', 'status': po.status},
                            status=status.HTTP_200_OK)
        with transaction.atomic():
            for it in po.items.all():
                stock, _ = Stock.objects.get_or_create(
                    product=it.product, branch=po.branch, defaults={'quantity': 0}
                )
                stock.quantity += it.quantity
                stock.save()
            po.status = 'received'
            po.save()
        return Response({'detail': 'Goods received — stock updated.', 'status': po.status},
                        status=status.HTTP_200_OK)

    @action(detail=True, methods=['GET'])
    def pdf(self, request, pk=None):
        po = self.get_object()
        # Reuse the shared branded PDF document (same layout as invoice/quotation)
        from apps.sales.utils import render_to_pdf
        from apps.sales.views import _company_ctx, _money_breakdown, _person, _amount_in_words
        company, currency, tax_rate = _company_ctx()
        items = [{
            'code': getattr(it.product, 'sku', '') or '',
            'description': it.product.name,
            'qty': it.quantity,
            'uom': it.get_unit_display(),
            'price': it.unit_cost,
            'total': it.total_cost,
        } for it in po.items.all()]
        subtotal_ex, tax_amount, total = _money_breakdown(po.total_amount, tax_rate)
        sup = po.supplier
        ctx = {
            'doc': {
                'type': 'PURCHASE ORDER', 'number_label': 'P.O. No.',
                'number': 'PO-%05d' % po.id, 'date': po.order_date or po.created_at,
                'valid_until': None, 'page': '1/1',
                'recipient_label': 'Supplier',
                'branch': po.branch.name if po.branch else '',
                'contact': _person(po.created_by),
                'authorized_by': _person(po.created_by),
                'status': None, 'payment_term': 'As agreed',
                'delivery_label': '', 'authorised_block': True,
            },
            'company': company,
            'customer': {
                'name': sup.name if sup else '',
                'phone': (sup.phone if sup else '') or (sup.contact_name if sup else ''),
                'address': sup.address if sup else '',
                'tin': '',
            },
            'items': items, 'currency': currency,
            'subtotal_ex': subtotal_ex, 'tax_rate': tax_rate,
            'tax_amount': tax_amount, 'total': total,
            'amount_words': _amount_in_words(total, currency),
        }
        return render_to_pdf('sales/pdf_document.html', ctx)

    @action(detail=True, methods=['POST'])
    def add_item(self, request, pk=None):
        po = self.get_object()
        product_id = request.data.get('product')
        quantity = int(request.data.get('quantity', 0))
        unit_cost = float(request.data.get('unit_cost', 0))
        
        product = Product.objects.get(pk=product_id)
        
        item = PurchaseOrderItem.objects.create(
            purchase_order=po,
            product=product,
            quantity=quantity,
            unit_cost=unit_cost,
            total_cost=quantity * unit_cost
        )
        
        # Update PO total
        po.total_amount = sum(i.total_cost for i in po.items.all())
        po.save()
        
        return Response(PurchaseOrderItemSerializer(item).data)

class TruckViewSet(viewsets.ModelViewSet):
    queryset = Truck.objects.all()
    serializer_class = TruckSerializer
    permission_classes = [permissions.IsAuthenticated, IsAfisaUgavi]

class TruckAllocationViewSet(viewsets.ModelViewSet):
    queryset = TruckAllocation.objects.all()
    serializer_class = TruckAllocationSerializer
    permission_classes = [permissions.IsAuthenticated, IsAfisaUgavi]

class GoodsReceivedNoteViewSet(viewsets.ModelViewSet):
    queryset = GoodsReceivedNote.objects.all()
    serializer_class = GoodsReceivedNoteSerializer
    # Stock Controller creates GRN (as per original requirements). Store Keeper verifies in UI.
    permission_classes = [permissions.IsAuthenticated, IsStockController]

    def perform_create(self, serializer):
        grn = serializer.save(created_by=self.request.user)
        
        # If successfully created, and PO is linked, update PO status to 'received'
        if grn.purchase_order:
            grn.purchase_order.status = 'received'
            grn.purchase_order.save()

    @action(detail=True, methods=['post'])
    def add_item(self, request, pk=None):
        grn = self.get_object()
        product_id = request.data.get('product')
        qty = request.data.get('quantity_received')
        
        try:
            product = Product.objects.get(id=product_id)
            item = GRNItem.objects.create(
                grn=grn,
                product=product,
                quantity_received=qty,
                remarks=request.data.get('remarks', '')
            )
            
            # UPDATE STOCK
            stock, created = Stock.objects.get_or_create(
                product=product, 
                branch=grn.branch,
                defaults={'quantity': 0}
            )
            stock.quantity += int(qty)
            stock.save()

            return Response(GRNItemSerializer(item).data)
        except Exception as e:
            return Response({'error': str(e)}, status=400)

class GRNListView(LoginRequiredMixin, TemplateView):
    template_name = 'inventory/grn_list.html'

class GRNCreateView(LoginRequiredMixin, TemplateView):
    template_name = 'inventory/grn_form.html'


class InventoryListView(LoginRequiredMixin, TemplateView):
    template_name = 'inventory/stock_list.html'

class BranchListView(LoginRequiredMixin, TemplateView):
    template_name = 'inventory/branch_list.html'

class BranchCreateView(LoginRequiredMixin, TemplateView):
    template_name = 'inventory/branch_create.html'

class SupplierListView(LoginRequiredMixin, TemplateView):
    template_name = 'inventory/supplier_list.html'

class PurchaseListView(LoginRequiredMixin, TemplateView):
    template_name = 'inventory/purchase_list.html'

class PurchaseCreateView(LoginRequiredMixin, TemplateView):
    template_name = 'inventory/purchase_create.html'

class RecentPurchaseListView(LoginRequiredMixin, TemplateView):
    template_name = 'inventory/recent_purchases.html'

class ProductListView(LoginRequiredMixin, TemplateView):
    template_name = 'product_list.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Products'
        context['resource'] = 'products'
        return context

class ProductCreateView(LoginRequiredMixin, TemplateView):
    template_name = 'product_list.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Create Product'
        context['resource'] = 'products'
        return context

class ProductImportView(LoginRequiredMixin, TemplateView):
    template_name = 'product_import.html'

class CategoryListView(LoginRequiredMixin, TemplateView):
    template_name = 'category_list.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Categories'
        context['resource'] = 'categories'
        return context

class ServicesListView(LoginRequiredMixin, TemplateView):
    template_name = 'product_list.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Services'
        context['resource'] = 'products?product_type=service'
        return context

class StockManagementView(LoginRequiredMixin, TemplateView):
    template_name = 'inventory_management.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Stock Management'
        context['resource'] = 'stocks'
        return context

class InventoryTransferView(LoginRequiredMixin, TemplateView):
    template_name = 'inventory_transfer.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Transfer'
        context['resource'] = 'transfers'
        return context

class InventoryHealthView(LoginRequiredMixin, TemplateView):
    template_name = 'inventory_health.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Inventory Health'
        context['resource'] = 'stocks'
        return context

class InventoryAgingView(LoginRequiredMixin, TemplateView):
    template_name = 'inventory_aging.html'
    def get_context_data(self, **kwargs):
        from apps.inventory.models import Stock, Purchase
        from django.utils import timezone
        context = super().get_context_data(**kwargs)
        
        stocks = Stock.objects.select_related('product', 'branch').all()
        now = timezone.now()
        
        aging_data = []
        for stock in stocks:
            # Find the latest purchase for this product in this branch
            last_purchase = Purchase.objects.filter(
                product=stock.product, 
                branch=stock.branch
            ).order_by('-date_purchased').first()
            
            last_date = last_purchase.date_purchased if last_purchase else stock.product.created_at
            days = (now - last_date).days
            
            aging_data.append({
                'product_name': stock.product.name,
                'branch_name': stock.branch.name,
                'quantity': stock.quantity,
                'last_date': last_date,
                'days': days,
                'status': 'danger' if days > 90 else 'warning' if days > 60 else 'info' if days > 30 else 'success'
            })
            
        context['aging_data'] = sorted(aging_data, key=lambda x: x['days'], reverse=True)
        context['title'] = 'Inventory Aging'
        context['resource'] = 'stocks'
        return context

class ABCAnalysisView(LoginRequiredMixin, TemplateView):
    template_name = 'inventory_abc.html'
    def get_context_data(self, **kwargs):
        from apps.sales.models import SaleItem
        from django.db.models import Sum, F
        context = super().get_context_data(**kwargs)
        
        # Calculate total revenue per product
        product_revenue = SaleItem.objects.values(
            'product__id', 'product__name'
        ).annotate(
            total_revenue=Sum(F('quantity') * F('price_at_sale'))
        ).order_by('-total_revenue')
        
        total_sum = sum(item['total_revenue'] for item in product_revenue) or 1
        
        cumulative_revenue = 0
        abc_data = []
        
        for item in product_revenue:
            cumulative_revenue += item['total_revenue']
            percentage = (cumulative_revenue / total_sum) * 100
            
            if percentage <= 70:
                abc_class = 'A'
                label = 'danger' # High importance
            elif percentage <= 90:
                abc_class = 'B'
                label = 'warning'
            else:
                abc_class = 'C'
                label = 'success'
                
            abc_data.append({
                'product_name': item['product__name'],
                'revenue': item['total_revenue'],
                'class': abc_class,
                'label': label
            })
            
        context['abc_data'] = abc_data
        context['title'] = 'ABC Analysis'
        context['resource'] = 'stocks'
        return context

class ProfitabilityReportView(LoginRequiredMixin, TemplateView):
    template_name = 'inventory_profitability.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Profitability Report'
        context['resource'] = 'stocks'
        return context

from django.contrib.auth.decorators import login_required

@login_required
def download_product_template(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="product_import_template.csv"'

    writer = csv.writer(response)
    # Headers
    writer.writerow(['Name', 'SKU', 'Category', 'Type', 'Cost', 'Price', 'Weight (kg)', 'Description', 'Opening Stock', 'Low Stock Alert'])
    # Sample Row
    writer.writerow(['Hammer', 'HMR-001', 'Tools', 'product', '15000', '25000', '1.5', 'Heavy duty steel hammer', '50', '5'])

    return response

class StockAdjustmentViewSet(viewsets.ModelViewSet):
    queryset = StockAdjustment.objects.all().order_by('-created_at')
    serializer_class = StockAdjustmentSerializer
    permission_classes = [permissions.DjangoModelPermissions]

    def perform_create(self, serializer):
        # Set the user to the current request user if any
        serializer.save(user=self.request.user if self.request.user.is_authenticated else None)

    def create(self, request, *args, **kwargs):
        payload = request.data.copy()
        serializer = self.get_serializer(data=payload)
        serializer.is_valid(raise_exception=True)
        
        with transaction.atomic():
            adjustment = serializer.save(user=request.user if request.user.is_authenticated else None)
            
            # Update Stock
            stock, _ = Stock.objects.get_or_create(
                product=adjustment.product, 
                branch=adjustment.branch,
                defaults={'quantity': 0}
            )
            
            if adjustment.adjustment_type == 'addition':
                stock.quantity += adjustment.quantity
            elif adjustment.adjustment_type == 'deduction':
                stock.quantity -= adjustment.quantity
            elif adjustment.adjustment_type == 'correction':
                stock.quantity = adjustment.quantity
            
            stock.save()

        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
    
class PurchaseOrderListView(LoginRequiredMixin, TemplateView):
    template_name = 'inventory/purchase_order_list.html'

class PurchaseOrderCreateView(LoginRequiredMixin, TemplateView):
    template_name = 'inventory/purchase_order_form.html'

class TruckListView(LoginRequiredMixin, TemplateView):
    template_name = 'inventory/truck_list.html'

class StockAdjustmentView(LoginRequiredMixin, TemplateView):
    template_name = 'inventory/stock_adjustment.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Stock Adjustment'
        context['resource'] = 'stocks'
        return context

class DriverViewSet(viewsets.ModelViewSet):
    queryset = Driver.objects.all()
    serializer_class = DriverSerializer
    permission_classes = [permissions.IsAuthenticated, IsStoreManager]

class TruckMaintenanceViewSet(viewsets.ModelViewSet):
    queryset = TruckMaintenance.objects.all().order_by('-date')
    serializer_class = TruckMaintenanceSerializer
    permission_classes = [permissions.IsAuthenticated, IsStoreManager]

    def perform_create(self, serializer):
        serializer.save(recorded_by=self.request.user)

class DriverListView(LoginRequiredMixin, TemplateView):
    template_name = 'inventory/driver_list.html'

class TruckMaintenanceView(LoginRequiredMixin, TemplateView):
    template_name = 'inventory/truck_maintenance.html'


# ----------------------------------------------------------------------------
# Purchases Report (filterable) + Excel / PDF export
# Built for the purchases manager (Afisa Ugavi); mirrors the finance
# expense report so the look-and-feel stays consistent.
# ----------------------------------------------------------------------------

def _filter_purchases(params):
    """Filter the Purchase queryset from request GET params.

    Supported filters (all optional): date_from, date_to, supplier, branch,
    q (product name search).
    """
    qs = (Purchase.objects.select_related('supplier', 'branch', 'product')
          .order_by('-date_purchased'))
    date_from = (params.get('date_from') or '').strip()
    date_to = (params.get('date_to') or '').strip()
    supplier = (params.get('supplier') or '').strip()
    branch = (params.get('branch') or '').strip()
    q = (params.get('q') or '').strip()

    if date_from:
        qs = qs.filter(date_purchased__date__gte=date_from)
    if date_to:
        qs = qs.filter(date_purchased__date__lte=date_to)
    if supplier:
        qs = qs.filter(supplier_id=supplier)
    if branch:
        qs = qs.filter(branch_id=branch)
    if q:
        qs = qs.filter(product__name__icontains=q)
    return qs


def _active_purchase_filter_labels(params):
    """Human-readable summary of the applied filters, for report headers."""
    labels = []
    if params.get('date_from') or params.get('date_to'):
        labels.append("Period: %s to %s" % (params.get('date_from') or 'start', params.get('date_to') or 'today'))
    if params.get('supplier'):
        s = Supplier.objects.filter(id=params.get('supplier')).first()
        if s:
            labels.append("Supplier: %s" % s.name)
    if params.get('branch'):
        br = Branch.objects.filter(id=params.get('branch')).first()
        if br:
            labels.append("Branch: %s" % br.name)
    if params.get('q'):
        labels.append('Product: "%s"' % params.get('q'))
    return labels or ["All purchases (no filters)"]


def _purchase_company_name():
    try:
        from apps.core.models import SystemSettings
        s = SystemSettings.objects.first()
        if s and getattr(s, 'company_name', None):
            return s.company_name
    except Exception:
        pass
    return "Umoja Hardware"


def _purchase_rows(qs):
    """Common row data for both exporters."""
    for p in qs:
        yield [
            p.date_purchased.strftime('%Y-%m-%d') if p.date_purchased else '',
            p.supplier.name if p.supplier else 'Unknown',
            p.product.name if p.product else '',
            p.branch.name if p.branch else '',
            int(p.quantity or 0),
            float(p.unit_cost or 0),
            float(p.total_cost or 0),
        ]


class PurchaseReportView(LoginRequiredMixin, TemplateView):
    template_name = 'inventory/purchase_report.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        params = self.request.GET
        qs = _filter_purchases(params)
        totals = qs.aggregate(amount=Sum('total_cost'), qty=Sum('quantity'))

        from urllib.parse import urlencode
        clean = {k: v for k, v in params.items() if k != 'format' and v}

        ctx.update({
            'purchases': qs,
            'total': totals['amount'] or Decimal('0'),
            'total_qty': totals['qty'] or 0,
            'count': qs.count(),
            'suppliers': Supplier.objects.all().order_by('name'),
            'branches': Branch.objects.all().order_by('name'),
            'filter_querystring': urlencode(clean),
            'f': params,  # echo back selected filter values into the form
        })
        return ctx


def _export_purchases_excel(qs, params):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Purchases Report"

    headers = ['Date', 'Supplier', 'Product', 'Branch', 'Qty', 'Unit Cost (TZS)', 'Total (TZS)']
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style='thin', color='DDDDDD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws['A1'] = _purchase_company_name()
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Purchases Report"
    ws['A2'].font = Font(bold=True, size=12, color="555555")
    row = 3
    for line in _active_purchase_filter_labels(params):
        ws.cell(row=row, column=1, value=line).font = Font(italic=True, color="666666")
        row += 1
    row += 1

    header_row = row
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.border = border
        c.alignment = Alignment(horizontal='center')

    total = Decimal('0')
    total_qty = 0
    r = header_row + 1
    for data in _purchase_rows(qs):
        for col, val in enumerate(data, start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.border = border
            if col in (5, 6, 7):
                c.alignment = Alignment(horizontal='right')
                if col in (6, 7):
                    c.number_format = '#,##0.00'
        total += Decimal(str(data[6]))
        total_qty += int(data[4])
        r += 1

    ws.cell(row=r, column=4, value="TOTAL").font = bold
    qc = ws.cell(row=r, column=5, value=total_qty)
    qc.font = bold
    qc.alignment = Alignment(horizontal='right')
    tc = ws.cell(row=r, column=7, value=float(total))
    tc.font = bold
    tc.number_format = '#,##0.00'
    tc.alignment = Alignment(horizontal='right')

    widths = [14, 26, 34, 18, 10, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="purchases_report_%s.xlsx"' % date.today().isoformat()
    wb.save(resp)
    return resp


def _export_purchases_pdf(qs, params):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=14 * mm, rightMargin=14 * mm,
                            topMargin=14 * mm, bottomMargin=14 * mm,
                            title="Purchases Report")
    styles = getSampleStyleSheet()
    cell = ParagraphStyle('cell', parent=styles['Normal'], fontSize=8, leading=10)
    elements = []
    elements.append(Paragraph(_purchase_company_name(), styles['Title']))
    elements.append(Paragraph("Purchases Report", styles['Heading2']))
    for line in _active_purchase_filter_labels(params):
        elements.append(Paragraph(line, styles['Italic']))
    elements.append(Spacer(1, 8))

    data = [['Date', 'Supplier', 'Product', 'Branch', 'Qty', 'Unit Cost', 'Total (TZS)']]
    total = Decimal('0')
    total_qty = 0
    for r in _purchase_rows(qs):
        data.append([
            r[0],
            Paragraph(str(r[1])[:120], cell),
            Paragraph(str(r[2])[:160], cell),
            r[3],
            str(r[4]),
            '{:,.2f}'.format(r[5]),
            '{:,.2f}'.format(r[6]),
        ])
        total += Decimal(str(r[6]))
        total_qty += int(r[4])
    data.append(['', '', '', 'TOTAL', str(total_qty), '', '{:,.2f}'.format(total)])

    table = Table(data, repeatRows=1, colWidths=[22 * mm, 50 * mm, 70 * mm, 30 * mm, 16 * mm, 28 * mm, 30 * mm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (4, 0), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#DDDDDD')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F5F7FA')]),
        ('FONTNAME', (3, -1), (-1, -1), 'Helvetica-Bold'),
        ('LINEABOVE', (0, -1), (-1, -1), 0.6, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)
    doc.build(elements)

    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = 'attachment; filename="purchases_report_%s.pdf"' % date.today().isoformat()
    return resp


class PurchaseReportExportView(LoginRequiredMixin, TemplateView):
    """GET ?format=excel|pdf plus the same filter params as the report page."""

    def get(self, request, *args, **kwargs):
        qs = _filter_purchases(request.GET)
        fmt = (request.GET.get('format') or 'excel').lower()
        if fmt == 'pdf':
            return _export_purchases_pdf(qs, request.GET)
        return _export_purchases_excel(qs, request.GET)
