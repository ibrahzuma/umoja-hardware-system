import io
from decimal import Decimal
from datetime import date
from rest_framework import viewsets, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Sum
from django.http import HttpResponse, Http404
from apps.users.permissions import IsAccountant, CanRecordSupplierPayment
from apps.sales.models import Sale
from apps.inventory.models import Branch
from .models import Expense, ExpenseCategory
from django.shortcuts import render
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from .forms import ExpenseForm
from .models import Expense, ExpenseCategory, Income, SupplierPayment, TaxPayment, PaymentReceipt, BankAccount
from .serializers import ExpenseSerializer, ExpenseCategorySerializer, IncomeSerializer, SupplierPaymentSerializer, TaxPaymentSerializer, PaymentReceiptSerializer, BankAccountSerializer

class ExpenseListView(LoginRequiredMixin, TemplateView):
    template_name = 'finance/expense_list.html'

class ExpenseCreateView(LoginRequiredMixin, CreateView):
    model = Expense
    form_class = ExpenseForm
    template_name = 'finance/add_expense.html'
    success_url = reverse_lazy('finance:expenses_list')

    def form_valid(self, form):
        if hasattr(self.request.user, 'branch'):
            form.instance.branch = self.request.user.branch
        form.instance.created_by = self.request.user
        return super().form_valid(form)

class RecentExpenseListView(LoginRequiredMixin, TemplateView):
    template_name = 'finance/recent_expenses.html'

class ExpenseCategoryListView(LoginRequiredMixin, TemplateView):
    template_name = 'finance/category_list.html'

class ExpenseCategoryViewSet(viewsets.ModelViewSet):
    queryset = ExpenseCategory.objects.all()
    serializer_class = ExpenseCategorySerializer
    permission_classes = [permissions.DjangoModelPermissions]

class BankAccountViewSet(viewsets.ModelViewSet):
    queryset = BankAccount.objects.all().order_by('name')
    serializer_class = BankAccountSerializer
    permission_classes = [permissions.DjangoModelPermissions]

class BankListView(LoginRequiredMixin, TemplateView):
    template_name = 'finance/bank_list.html'

class ExpenseViewSet(viewsets.ModelViewSet):
    queryset = Expense.objects.select_related('category', 'bank', 'branch').order_by('-date_incurred')
    serializer_class = ExpenseSerializer
    permission_classes = [permissions.DjangoModelPermissions]

class IncomeListView(LoginRequiredMixin, TemplateView):
    template_name = 'finance/income_list.html'

class IncomeViewSet(viewsets.ModelViewSet):
    queryset = Income.objects.all().order_by('-date_received')
    serializer_class = IncomeSerializer
    permission_classes = [permissions.DjangoModelPermissions]

class SupplierPaymentViewSet(viewsets.ModelViewSet):
    queryset = SupplierPayment.objects.all().order_by('-payment_date')
    serializer_class = SupplierPaymentSerializer
    permission_classes = [permissions.IsAuthenticated, CanRecordSupplierPayment]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

class SupplierPaymentListView(LoginRequiredMixin, TemplateView):
    template_name = 'finance/supplier_payment_list.html'

class TaxPaymentViewSet(viewsets.ModelViewSet):
    queryset = TaxPayment.objects.all().order_by('-payment_date')
    serializer_class = TaxPaymentSerializer
    permission_classes = [permissions.IsAuthenticated, IsAccountant]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

class TaxPaymentListView(LoginRequiredMixin, TemplateView):
    template_name = 'finance/tax_payment_list.html'

class DebtorListView(LoginRequiredMixin, TemplateView):
    template_name = 'finance/debtor_list.html'


class PaymentReceiptViewSet(viewsets.ModelViewSet):
    queryset = PaymentReceipt.objects.select_related(
        'sale', 'customer', 'issued_by', 'created_by'
    ).order_by('-created_at')
    serializer_class = PaymentReceiptSerializer
    permission_classes = [permissions.IsAuthenticated, IsAccountant]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=False, methods=['get'])
    def lookup(self, request):
        """Preview an invoice before a receipt is logged against it.

        GET /api/payment-receipts/lookup/?invoice=INV-123
        Returns the invoice total, customer, issuer, and how much has already
        been received via prior receipts, so finance sees the outstanding
        balance before saving.
        """
        invoice = (request.query_params.get('invoice') or '').strip()
        if not invoice:
            return Response({'found': False, 'detail': 'No invoice number provided.'})

        sale = Sale.objects.select_related('customer', 'user').filter(invoice_number=invoice).first()
        if sale is None:
            return Response({'found': False, 'detail': 'No invoice found with that number.'})

        invoice_amount = sale.total_amount or Decimal('0')
        already_paid = sale.payment_receipts.aggregate(total=Sum('amount_paid'))['total'] or Decimal('0')
        issuer = sale.user
        return Response({
            'found': True,
            'invoice_number': sale.invoice_number,
            'customer_name': sale.customer.name if sale.customer else (sale.customer_name or ''),
            'issued_by_name': (issuer.get_full_name() or issuer.username) if issuer else '',
            'invoice_amount': str(invoice_amount),
            'already_paid': str(already_paid),
            'outstanding': str(invoice_amount - already_paid),
        })


class PaymentReceiptListView(LoginRequiredMixin, TemplateView):
    template_name = 'finance/payment_receipt_list.html'


# ----------------------------------------------------------------------------
# Expense Report (filterable) + Excel / PDF export
# ----------------------------------------------------------------------------

def _filter_expenses(params):
    """Filter the expense queryset from request GET params.

    Supported filters (all optional): date_from, date_to, category, bank
    ('none' = cash/no bank), branch, q (description search).
    """
    qs = Expense.objects.select_related('category', 'bank', 'branch').order_by('-date_incurred')
    date_from = (params.get('date_from') or '').strip()
    date_to = (params.get('date_to') or '').strip()
    category = (params.get('category') or '').strip()
    bank = (params.get('bank') or '').strip()
    branch = (params.get('branch') or '').strip()
    q = (params.get('q') or '').strip()

    if date_from:
        qs = qs.filter(date_incurred__gte=date_from)
    if date_to:
        qs = qs.filter(date_incurred__lte=date_to)
    if category:
        qs = qs.filter(category_id=category)
    if bank:
        if bank == 'none':
            qs = qs.filter(bank__isnull=True)
        else:
            qs = qs.filter(bank_id=bank)
    if branch:
        qs = qs.filter(branch_id=branch)
    if q:
        qs = qs.filter(description__icontains=q)
    return qs


def _active_filter_labels(params):
    """Human-readable summary of the applied filters, for report headers."""
    labels = []
    if params.get('date_from') or params.get('date_to'):
        labels.append("Period: %s to %s" % (params.get('date_from') or 'start', params.get('date_to') or 'today'))
    if params.get('category'):
        cat = ExpenseCategory.objects.filter(id=params.get('category')).first()
        if cat:
            labels.append("Category: %s" % cat.name)
    if params.get('bank'):
        if params.get('bank') == 'none':
            labels.append("Bank: Cash / none")
        else:
            b = BankAccount.objects.filter(id=params.get('bank')).first()
            if b:
                labels.append("Bank: %s" % b.name)
    if params.get('branch'):
        br = Branch.objects.filter(id=params.get('branch')).first()
        if br:
            labels.append("Branch: %s" % br.name)
    if params.get('q'):
        labels.append('Search: "%s"' % params.get('q'))
    return labels or ["All expenses (no filters)"]


def _company_name():
    try:
        from apps.core.models import SystemSettings
        s = SystemSettings.objects.first()
        if s and getattr(s, 'company_name', None):
            return s.company_name
    except Exception:
        pass
    return "Umoja Hardware"


class ExpenseReportView(LoginRequiredMixin, TemplateView):
    template_name = 'finance/expense_report.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        params = self.request.GET
        qs = _filter_expenses(params)
        total = qs.aggregate(t=Sum('amount'))['t'] or Decimal('0')

        # querystring carrying the current filters (minus export format) for the buttons
        from urllib.parse import urlencode
        clean = {k: v for k, v in params.items() if k != 'format' and v}

        ctx.update({
            'expenses': qs,
            'total': total,
            'count': qs.count(),
            'categories': ExpenseCategory.objects.all().order_by('name'),
            'banks': BankAccount.objects.all().order_by('name'),
            'branches': Branch.objects.all().order_by('name'),
            'filter_querystring': urlencode(clean),
            'f': params,  # echo back selected filter values into the form
        })
        return ctx


def _expense_rows(qs):
    """Common row data for both exporters."""
    for e in qs:
        yield [
            e.date_incurred.strftime('%Y-%m-%d') if e.date_incurred else '',
            e.category.name if e.category else 'Uncategorized',
            e.description or '',
            e.branch.name if e.branch else '',
            e.bank.name if e.bank else 'Cash / none',
            float(e.amount or 0),
        ]


def _export_excel(qs, params):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Expense Report"

    headers = ['Date', 'Category', 'Description', 'Branch', 'Bank', 'Amount (TZS)']
    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style='thin', color='DDDDDD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws['A1'] = _company_name()
    ws['A1'].font = title_font
    ws['A2'] = "Expense Report"
    ws['A2'].font = Font(bold=True, size=12, color="555555")
    row = 3
    for line in _active_filter_labels(params):
        ws.cell(row=row, column=1, value=line).font = Font(italic=True, color="666666")
        row += 1
    row += 1

    header_row = row
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.border = border
        c.alignment = Alignment(horizontal='center')

    total = Decimal('0')
    r = header_row + 1
    for data in _export_rows_for_excel(qs):
        for col, val in enumerate(data, start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.border = border
            if col == 6:
                c.number_format = '#,##0'
                c.alignment = Alignment(horizontal='right')
        total += Decimal(str(data[5]))
        r += 1

    ws.cell(row=r, column=5, value="TOTAL").font = bold
    tc = ws.cell(row=r, column=6, value=float(total))
    tc.font = bold
    tc.number_format = '#,##0'
    tc.alignment = Alignment(horizontal='right')

    widths = [14, 20, 40, 18, 22, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="expense_report_%s.xlsx"' % date.today().isoformat()
    wb.save(resp)
    return resp


def _export_rows_for_excel(qs):
    return list(_expense_rows(qs))


def _export_pdf(qs, params):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=14 * mm, rightMargin=14 * mm,
                            topMargin=14 * mm, bottomMargin=14 * mm,
                            title="Expense Report")
    styles = getSampleStyleSheet()
    cell = ParagraphStyle('cell', parent=styles['Normal'], fontSize=8, leading=10)
    elements = []
    elements.append(Paragraph(_company_name(), styles['Title']))
    elements.append(Paragraph("Expense Report", styles['Heading2']))
    for line in _active_filter_labels(params):
        elements.append(Paragraph(line, styles['Italic']))
    elements.append(Spacer(1, 8))

    data = [['Date', 'Category', 'Description', 'Branch', 'Bank', 'Amount (TZS)']]
    total = Decimal('0')
    for r in _expense_rows(qs):
        data.append([
            r[0], r[1],
            Paragraph(str(r[2])[:200], cell),
            r[3], r[4],
            '{:,.0f}'.format(r[5]),
        ])
        total += Decimal(str(r[5]))
    data.append(['', '', '', '', 'TOTAL', '{:,.0f}'.format(total)])

    table = Table(data, repeatRows=1, colWidths=[22 * mm, 30 * mm, 90 * mm, 30 * mm, 35 * mm, 30 * mm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (5, 0), (5, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#DDDDDD')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F5F7FA')]),
        ('FONTNAME', (4, -1), (-1, -1), 'Helvetica-Bold'),
        ('LINEABOVE', (0, -1), (-1, -1), 0.6, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)
    doc.build(elements)

    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = 'attachment; filename="expense_report_%s.pdf"' % date.today().isoformat()
    return resp


class ExpenseReportExportView(LoginRequiredMixin, TemplateView):
    """GET ?format=excel|pdf plus the same filter params as the report page."""

    def get(self, request, *args, **kwargs):
        qs = _filter_expenses(request.GET)
        fmt = (request.GET.get('format') or 'excel').lower()
        if fmt in ('xlsx', 'excel'):
            return _export_excel(qs, request.GET)
        if fmt == 'pdf':
            return _export_pdf(qs, request.GET)
        raise Http404("Unknown export format")
